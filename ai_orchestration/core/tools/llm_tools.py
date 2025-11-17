"""
LLM Tools - Functions that call LLMs as tools
Uses @tool decorator for auto-registration
"""

from typing import Dict
from pathlib import Path
import sys
import base64
from pypdf import PdfReader
import openpyxl

# Import ask_llm API (renamed to avoid collision with tools)
from ai_orchestration.third_pary_llms.ask_llm import ask_llm as _call_llm_api

from ..modals import WorkspaceModal
from ..tool_decorator import tool


@tool(execution_type="function")
def ask_ai(workspace: WorkspaceModal, task_id: str, agent_model: str, question: str, context: str = "") -> Dict:
    """ask_ai"""
    # MOP: agent_model injected from agent.force_model
    # Build messages
    messages = []
    if context:
        messages.append({
            "role": "user",
            "content": f"Context: {context}\n\nQuestion: {question}"
        })
    else:
        messages.append({
            "role": "user",
            "content": question
        })

    # Call LLM API with agent's model
    response = _call_llm_api(
        messages=messages,
        model=agent_model
    )

    # Extract answer
    if "error" in response:
        return {
            **response,
            "model": agent_model
        }

    # Get text content from response
    content = response.get("content", [])
    answer = ""
    for block in content:
        if isinstance(block, dict) and block.get("type") == "text":
            answer += block.get("text", "")
        elif isinstance(block, str):
            answer += block

    return {
        "answer": answer,
        "model": agent_model
    }


@tool(execution_type="function")
def ask_data(workspace: WorkspaceModal, task_id: str, agent_model: str, file_path: str = None, file_paths: list = None, question: str = "") -> Dict:
    """ask_data"""
    # MOP: agent_model injected from agent.force_model
    # Handle both single file and multiple files
    if file_paths is None:
        if file_path is None:
            return {"error": "Must provide file_path or file_paths"}
        paths = [file_path]
    else:
        paths = file_paths if isinstance(file_paths, list) else [file_paths]

    # Build content blocks for message
    content_blocks = []
    file_types = []
    image_exts = {'.png', '.jpg', '.jpeg', '.gif', '.webp', '.bmp'}

    # Process each file
    for fpath in paths:
        file_ext = Path(fpath).suffix.lower()

        # For binary files (PDF, Excel, images), need workspace_path
        binary_exts = {'.pdf', '.xlsx', '.xls', '.png', '.jpg', '.jpeg', '.gif', '.webp', '.bmp'}
        if file_ext in binary_exts:
            # Binary files require workspace_path (offensive - crash if not set)
            actual_path = workspace.workspace_path / fpath
            if not actual_path.exists():
                return {
                    "error": f"File not found on disk: {actual_path}",
                    "file_path": fpath
                }
            # Use file extension to determine type
            type_map = {
                '.pdf': 'pdf', '.xlsx': 'excel', '.xls': 'excel',
                '.png': 'image', '.jpg': 'image', '.jpeg': 'image',
                '.gif': 'image', '.webp': 'image', '.bmp': 'image'
            }
            file_type = type_map.get(file_ext, 'binary')
            file_types.append(file_type)
        else:
            # Text files - read from document registry
            doc = workspace.get_document(fpath)
            if not doc:
                return {
                    "error": f"File not found in workspace: {fpath}",
                    "file_path": fpath
                }
            file_types.append(doc.doc_type)
            # For text files, set actual_path only if workspace_path exists (for fallback reads)
            actual_path = (workspace.workspace_path / fpath) if workspace.workspace_path else None

        # Image files - add vision content
        if file_ext in image_exts:
            # Images are binary files, actual_path already set above
            if not actual_path.exists():
                return {
                    "error": f"Image file not found on disk: {actual_path}",
                    "file_path": fpath
                }

            # Read image as base64
            image_data = actual_path.read_bytes()
            image_b64 = base64.b64encode(image_data).decode('utf-8')

            # Determine media type
            media_type_map = {
                '.png': 'image/png',
                '.jpg': 'image/jpeg',
                '.jpeg': 'image/jpeg',
                '.gif': 'image/gif',
                '.webp': 'image/webp'
            }
            media_type = media_type_map.get(file_ext, 'image/jpeg')

            # Add image block
            content_blocks.append({
                "type": "image",
                "source": {
                    "type": "base64",
                    "media_type": media_type,
                    "data": image_b64
                }
            })

        # PDF files - extract text
        elif file_ext == '.pdf':
            if not actual_path.exists():
                return {
                    "error": f"PDF file not found on disk: {actual_path}",
                    "file_path": fpath
                }

            # Extract text from all pages
            reader = PdfReader(str(actual_path))
            pdf_text = ""
            for page_num, page in enumerate(reader.pages, 1):
                text = page.extract_text()
                if text:
                    pdf_text += f"Page {page_num}:\n{text}\n\n"

            content_blocks.append({
                "type": "text",
                "text": f"File: {fpath} (PDF - {len(reader.pages)} pages)\n\n{pdf_text}\n"
            })

        # Excel files - extract data
        elif file_ext in {'.xlsx', '.xls'}:
            if not actual_path.exists():
                return {
                    "error": f"Excel file not found on disk: {actual_path}",
                    "file_path": fpath
                }

            # Load workbook and extract all sheets
            wb = openpyxl.load_workbook(str(actual_path))
            excel_text = ""
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                excel_text += f"Sheet: {sheet_name}\n"

                # Convert to CSV-like format
                for row in sheet.iter_rows(values_only=True):
                    row_text = ", ".join([str(cell) if cell is not None else "" for cell in row])
                    excel_text += row_text + "\n"
                excel_text += "\n"

            content_blocks.append({
                "type": "text",
                "text": f"File: {fpath} (Excel - {len(wb.sheetnames)} sheets)\n\n{excel_text}\n"
            })

        # Python/JavaScript files - enhanced code analysis
        elif file_ext in {'.py', '.js', '.jsx', '.ts', '.tsx'}:
            # Read from document registry
            doc = workspace.get_document(fpath)
            if not doc:
                # Fallback: read directly from file
                file_content = actual_path.read_text(encoding='utf-8')
            else:
                file_content = doc.render_from_json()

            # Determine language
            lang_map = {'.py': 'Python', '.js': 'JavaScript', '.jsx': 'JavaScript (JSX)',
                       '.ts': 'TypeScript', '.tsx': 'TypeScript (TSX)'}
            language = lang_map.get(file_ext, 'Code')

            # Extract code structure info
            lines = file_content.split('\n')
            imports = []
            functions = []
            classes = []

            if file_ext == '.py':
                for line in lines:
                    stripped = line.strip()
                    if stripped.startswith('import ') or stripped.startswith('from '):
                        imports.append(stripped)
                    elif stripped.startswith('def '):
                        func_name = stripped.split('(')[0].replace('def ', '')
                        functions.append(func_name)
                    elif stripped.startswith('class '):
                        class_name = stripped.split('(')[0].split(':')[0].replace('class ', '')
                        classes.append(class_name)

            elif file_ext in {'.js', '.jsx', '.ts', '.tsx'}:
                for line in lines:
                    stripped = line.strip()
                    if stripped.startswith('import ') or stripped.startswith('export '):
                        imports.append(stripped)
                    elif 'function ' in stripped or '=>' in stripped:
                        if 'function ' in stripped:
                            parts = stripped.split('function ')
                            if len(parts) > 1:
                                func_name = parts[1].split('(')[0].strip()
                                if func_name:
                                    functions.append(func_name)
                    elif stripped.startswith('class '):
                        class_name = stripped.split('{')[0].split('extends')[0].replace('class ', '').strip()
                        classes.append(class_name)

            # Build enhanced context
            structure_info = f"File: {fpath} ({language})\n"
            structure_info += f"Lines: {len(lines)}\n"
            if imports:
                structure_info += f"Imports: {len(imports)}\n"
            if functions:
                structure_info += f"Functions: {', '.join(functions[:10])}\n"
            if classes:
                structure_info += f"Classes: {', '.join(classes)}\n"
            structure_info += f"\n{file_content}\n"

            content_blocks.append({
                "type": "text",
                "text": structure_info
            })

        else:
            # Text-based files - add text block
            doc = workspace.get_document(fpath)
            if not doc:
                # Fallback: read directly from file if it exists
                if actual_path.exists():
                    file_content = actual_path.read_text(encoding='utf-8')
                else:
                    return {
                        "error": f"File not found: {fpath}",
                        "file_path": fpath
                    }
            else:
                file_content = doc.render_from_json()

            content_blocks.append({
                "type": "text",
                "text": f"File: {fpath}\n\n{file_content}\n"
            })

    # Add question at the end
    content_blocks.append({
        "type": "text",
        "text": f"\nQuestion: {question}"
    })

    # Build message
    messages = [{
        "role": "user",
        "content": content_blocks
    }]

    # Call LLM API with agent's model
    response = _call_llm_api(
        messages=messages,
        model=agent_model
    )

    # Extract answer
    if "error" in response:
        return {
            **response,
            "file_paths": paths,
            "file_types": file_types
        }

    # Get text content from response
    content = response.get("content", [])
    answer = ""
    for block in content:
        if isinstance(block, dict) and block.get("type") == "text":
            answer += block.get("text", "")
        elif isinstance(block, str):
            answer += block

    # Return with both old and new format for backward compatibility
    result = {
        "answer": answer,
        "file_paths": paths,
        "file_types": file_types
    }

    # Backward compatibility: if single file, also include file_path and file_type
    if len(paths) == 1:
        result["file_path"] = paths[0]
        result["file_type"] = file_types[0]

    return result


# search_web REMOVED - no config file exists
